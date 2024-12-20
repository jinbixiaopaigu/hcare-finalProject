# -*- coding: utf-8 -*-
# @Author  : shaw-lee

from io import BytesIO
import os,socket,threading,re,base64,inspect,ipaddress,math,psutil
import typing_extensions
import time
from typing import Any, Callable, Dict, List, Literal, Optional, Tuple, Type, get_args, get_origin
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill,Alignment
from pydantic import BaseModel
from pydantic._internal import _typing_extra
from werkzeug.exceptions import NotFound
from werkzeug.datastructures import FileStorage
from flask import Response, request
from jwt import api_jwt

from owl_common.base.schema_excel import ExcelAccess

from ..constant import Constants


class UtilException(Exception):
    
    def __init__(self, message, status):
        super().__init__(message)
        self.status = status


class StringUtil:
        
    @classmethod
    def to_bool(cls, value) -> bool:
        """
        将值转换为布尔值

        Args:
            value (str or bytes): 输入

        Returns:
            bool: 转化后的布尔值
        """
        if isinstance(value, str):
            value = value.lower()
            if value in ['true', '1', 'yes', 'y']:
                return True
            elif value in ['false', '0', 'no', 'n']:
                return False
        elif isinstance(value, bytes):
            value = value.decode('utf-8')
            return cls.to_bool(value)
        elif value is None:
            return False
        else:
            raise TypeError('value must be str or bytes')
        return bool(value)
    
    @classmethod
    def to_int(cls, value) -> int:
        """
        将值转换为整数

        Args:
            value (str): 输入

        Returns:
            int: 转化后的整数
        """
        if isinstance(value, str):
            value = value.strip()
            if value.isdigit():
                return int(value)
        return int(value)
    
    @classmethod
    def to_float(cls, value) -> float:
        """
        将值转换为浮点数
        
        Args:
            value (str): 输入

        Returns:
            float: 转化后的浮点数
        """
        if isinstance(value, str):
            value = value.strip()
            if value.replace('.', '', 1).isdigit():
                return float(value)
        return float(value)
    
    @classmethod
    def to_str(cls, value) -> str:
        """
        将值转换为字符串

        Args:
            value (_type_): 输入

        Returns:
            str: 转化后的字符串
        """
        if isinstance(value, bool):
            return str(value).lower()
        elif isinstance(value, bytes):
            value = value.decode('utf-8')
            return value
        return str(value)
    
    @classmethod
    def to_datetime(cls, value) -> datetime:
        """
        将字符串转换为datetime类型  

        Args:
            value (str): 字符类型日期字符串

        Returns:
            datetime: datetime类型日期
        """
        if isinstance(value, str):
            value = value.strip()
            try:
                return datetime.strptime(value, cls.datatime_format)
            except ValueError:
                pass
        return value

    @classmethod
    def ishttp(cls, val:str) -> bool:
        """
        判断是否为http或https开头的url

        Args:
            val (str): 输入字符串

        Returns:
            bool: 是否为http或https开头的url
        """
        return val.startswith('http://') or val.startswith('https://')
    
    @classmethod
    def pad_left(cls, value, length:int) -> str:
        """
        左填充字符串，使其长度达到指定长度

        Args:
            value (_type_): 输入字符串
            length (int): 目标长度

        Returns:
            str: 填充后的字符串
        """
        return str(value).zfill(length)
    
    @classmethod
    def substring_after(cls, string:str, separator:str) -> str:
        """
        获取字符串string中第一个分隔符separator之后的字符串

        Args:
            string (str): 输入字符串
            separator (str): 分隔符

        Returns:
            str: 转换后的字符串
        """
        if separator in string:
            return string.split(separator, 1)[1]
        return ""
    
        
class DictUtil:
    
    @classmethod
    def upper_key(cls,dict_obj:dict) -> dict:
        '''
        将配置对象中的所有键转换为大写
        
        Args:
            dict_obj(dict): 字典
        
        Returns:
            dict: 转换后的字典
        '''
        new_configobj = {}
        for k in dict_obj.keys():
            v = dict_obj.get(k)
            if k.islower():
                if isinstance(v, dict):
                    new_configobj[k.upper()] = cls.upper_key(v)
                else:
                    new_configobj[k.upper()] = v
            else:
                new_configobj[k] = v
        return new_configobj
    
    @classmethod
    def lower_key(cls,dict_obj:dict) -> dict:
        '''
        将配置对象中的所有键转换为小写
        
        Args:
            dict_obj(dict): 字典
        
        Returns:
            dict: 转换后的字典
        '''
        new_configobj = {}
        for k in dict_obj.keys():
            v = dict_obj.get(k)
            if k.islower():
                if isinstance(v, dict):
                    new_configobj[k.upper()] = cls.upper_key(v)
                else:
                    new_configobj[k.upper()] = v
        return new_configobj
    
    @classmethod
    def flatten(cls,dict_obj) -> dict:
        '''
        将字典展平为一级字典
        
        Args:
            dict_obj(dict): 字典
        
        Returns:
            dict: 展平后的字典
        '''
        new_dict = {}
        inner_dict = {}
        for k,v in dict_obj.items():
            if isinstance(v,dict):
                inner_dict.update(cls.flatten(v))
            else:
                new_dict[k] = v
        new_dict.update(inner_dict)
        return new_dict
        
    @classmethod
    def format_value(cls,dict_obj) -> dict:
        '''
        格式化字典的值
        
        Args:
            dict_obj(dict): 字典
        
        Returns:
            dict: 格式化后的字典
        '''
        pattern = re.compile("\{(.*?)\}")
        new_dict = {}
        for k,v in dict_obj.items():
            if isinstance(v,str) and re.match(pattern,v):
                v_v = v.format(**dict_obj)
                new_dict[k] = v_v
            else:
                new_dict[k] = v
        return new_dict
    
    @classmethod
    def recurive_key(cls,dict_obj,pre_key="") -> dict:
        '''
        递归处理字典中的键
        
        Args:
            dict_obj(dict): 字典
            pre_key(str): 前缀
        
        Returns:
            dict: 处理后的字典
        '''
        new_dict = {}
        new_key = ""
        for k1,v1 in dict_obj.items():
            if pre_key=="":
                new_key = k1
            else:
                new_key = "{}.{}".format(pre_key,k1)
            if isinstance(v1,dict):
                next_dict = cls.recurive_key(v1,new_key)
                new_dict[new_key] = next_dict
            else:
                new_dict[new_key] = v1
        return new_dict
    

        
class Base64Util:
    
    @classmethod
    def decode(cls,data:str, is_padding=True) -> str:
        '''
        base64解码
        
        Args:
            data(str): base64编码数据
            is_padding(bool): 是否需要补位
        
        Returns:
            str: 解码后数据
        '''
        suplus = len(data) % 4
        if is_padding:
            missing_padding = 4 - suplus
            data += '='* missing_padding
        else:
            data = data[:-suplus] if suplus else data
        return str(base64.b64decode(data))
        

class TokenUtil:
    
    default_algorithm = "HS512"    
    
    default_headers = {
        "typ": None,
        "alg": default_algorithm
    }
    
    @classmethod
    def encode(cls, payload, secret, headers=None) -> str:
        '''
        编码生成jwt token
        
        Args:
            payload(dict): 载荷
            secret(str): 密钥
            headers(dict): 头部
        
        Returns:
            str: jwt token
        '''
        if headers is None:
            headers = cls.default_headers
        if "alg" not in headers:
            headers["alg"] = cls.default_algorithm
        else:
            if not headers["alg"]:
                headers["alg"] = cls.default_algorithm
        algorithm = headers["alg"]
        secret_decoded = Base64Util.decode(secret)
        jwt = api_jwt.encode(
            payload,
            secret_decoded,
            algorithm,
            headers=headers
        )
        return jwt
    
    @classmethod
    def decode(cls, jwt, secret, algorithms=None, verify=True) -> dict:
        '''
        解码jwt token
        
        Args:
            jwt(str): jwt token
            secret(str): 密钥
            algorithms(str): 算法
            verify(bool): 是否验证
        
        Returns:
            dict: 解码后的payload
        '''
        if algorithms is None:
            algorithms = cls.default_algorithm
        secret_decoded = Base64Util.decode(secret)
        payload = api_jwt.decode(jwt, secret_decoded, algorithms=algorithms, verify=verify)
        return payload

    @classmethod
    def get_from_request(cls) -> str:
        '''
        从请求头中获取token
        
        Returns:
            str: token
        '''
        authorization = request.headers.get('Authorization', None)
        if authorization is None:
            raise Exception('Authorization header not found')
        authorization_split = authorization.split()
        if len(authorization_split)!= 2 or authorization_split[0].lower()!= 'bearer':
            raise Exception('Invalid authorization header')
        return authorization_split[1]
    
    @classmethod
    def verify_from_request(cls, key, algorithms=None):
        '''
        从请求头中获取token，并验证token
        
        Args:
            key(str): 密钥
            algorithms(str): 算法
        
        Raises:
            UtilException: 验证失败
        '''
        encoded_token = cls.get_token_from_request()
        cls.decode(encoded_token, key, algorithms=algorithms, verify=True)
        

class IpUtil:
    
    @classmethod
    def get_ip(cls):
        '''
        获取请求ip
        
        Returns:
            str: ip
        '''
        ip = None
        if 'HTTP_X_FORWARDED_FOR' in request.headers:
            ip = request.headers['HTTP_X_FORWARDED_FOR']
        elif 'REMOTE_ADDR' in request.headers:
            ip = request.headers['REMOTE_ADDR']
        ip, _ = request.host.rsplit(':', 1)
        ip = "127.0.0.1" if ip == "localhost" else ip
        return ip
    
    @classmethod
    def get_local_ips(cls) -> List[str]:
        '''
        获取本地ip
        
        Returns:
            List[str]: 本地ip列表
        '''
        ips = []
        addrs = psutil.net_if_addrs()  # 获取所有网络接口信息
        for interface, addr_list in addrs.items():
            for addr in addr_list:
                if addr.family == socket.AF_INET and not addr.address.startswith("127."):
                    ips.append(addr.address)
        return ips

    def is_valid_ip(ip: str) -> bool:
        try:
            ipaddress.ip_address(ip)
            return True
        except ValueError:
            return False

class AddressUtil:
    
    @classmethod
    def get_address(cls, ip) -> str:
        '''
        根据ip获取地址
        
        Args:
            ip(str): ip
        
        Returns:
            str: 地址
        '''
        # todo
        address = None
        return address


class UserAgentUtil:
    
    @classmethod
    def get_user_agent(cls) -> str:
        '''
        获取请求头中的user-agent
        
        Returns:
            str: user-agent
        '''
        user_agent = request.headers.get('User-Agent', None)
        return user_agent
    
    @classmethod
    def is_mobile(cls) -> bool:
        '''
        判断是否为移动端
        
        Returns:
            bool: 是否为移动端
        '''
        user_agent = cls.get_user_agent()
        if user_agent is None:
            return False
        mobile_agents = [
            'Android', 'iPhone', 'SymbianOS', 'Windows Phone', 'iPad', 'iPod'
        ]
        for agent in mobile_agents:
            if agent in user_agent:
                return True
        return False
    
    @classmethod
    def browser(cls) -> Literal['Chrome', 'Firefox', 'Safari', 'IE', None]:
        '''
        获取浏览器类型
        
        Returns:
            Literal['Chrome', 'Firefox', 'Safari', 'IE', None]: 浏览器类型
        '''
        user_agent = cls.get_user_agent()
        if user_agent is None:
            return None
        if 'Chrome' in user_agent:
            return 'Chrome'
        elif 'Firefox' in user_agent:
            return 'Firefox'
        elif 'Safari' in user_agent:
            return 'Safari'
        elif 'MSIE' in user_agent or 'Trident' in user_agent:
            return 'IE'
        else:
            return None
    
    @classmethod
    def os(cls) -> Literal['Windows', 'Mac', 'Linux', 'Unix', None]:
        '''
        获取操作系统类型
        
        Returns:
            Literal['Windows', 'Mac', 'Linux', 'Unix', None]: 操作系统类型
        '''
        user_agent = cls.get_user_agent()
        if user_agent is None:
            return None
        if 'Windows NT' in user_agent:
            return 'Windows'
        elif 'Macintosh' in user_agent:
            return 'Mac'
        elif 'Linux' in user_agent:
            return 'Linux'
        elif 'Unix' in user_agent:
            return 'Unix'
        else:
            return None
    
    @classmethod
    def is_pc(cls) -> bool:
        '''
        判断是否为PC端
        
        Returns:
            bool: 是否为PC端
        '''
        user_agent = cls.get_user_agent()
        if user_agent is None:
            return False
        pc_agents = [
            'Windows NT', 'Macintosh', 'Linux', 'Unix', 'FreeBSD', 'OpenBSD',
            'NetBSD', 'SunOS', 'AIX', 'HP-UX', 'IRIX', 'OSF1', ' SCO', 'IRIX64'
        ]
        for agent in pc_agents:
            if agent in user_agent:
                return True
        return False
    
    @classmethod
    def is_weixin(cls) -> bool:
        '''
        判断是否为微信端
        
        Returns:
            bool: 是否为微信端
        '''
        user_agent = cls.get_user_agent()
        if user_agent is None:
            return False
        weixin_agents = [
            'MicroMessenger', 'WeChat', 'QQ', 'Weibo', 'TencentTraveler',
            'QQBrowser', 'QQMobile', 'QQScan', 'Tenvideo', 'SogouExplorer',
        ]
        pass
    

class MimeTypeUtil:
    
    IMAGE_PNG = "image/png"
    
    IMAGE_JPG = "image/jpg"
    
    IMAGE_JPEG = "image/jpeg"
    
    IMAGE_BMP = "image/bmp"
    
    IMAGE_GIF = "image/gif"
    
    VIDEO_EXTENSION = [ "mp4", "avi", "rmvb" ]
    
    MEDIA_EXTENSION = [ "swf", "flv", "mp3", "wav", "wma", "wmv", "mid", "avi", "mpg",
            "asf", "rm", "rmvb" ]
    
    FLASH_EXTENSION = [ "swf", "flv" ]
    
    IMAGE_EXTENSION = [ "bmp", "gif", "jpg", "jpeg", "png" ]
    
    DEFAULT_ALLOWED_EXTENSION = [
        # 图片
        "bmp", "gif", "jpg", "jpeg", "png",
        # word excel powerpoint
        "doc", "docx", "xls", "xlsx", "ppt", "pptx", "html", "htm", "txt",
        # 压缩文件
        "rar", "zip", "gz", "bz2",
        # 视频格式
        "mp4", "avi", "rmvb",
        # pdf
        "pdf" ]
    
    def get_extension(cls, mime_type:str):
        '''
        根据mime_type获取文件扩展名
        
        Args:
            mime_type(str): mime_type
        
        Returns:
            str: 文件扩展名
        '''
        match mime_type:
            case cls.IMAGE_PNG:
                return "png"
            case cls.IMAGE_JPG:
                return "jpg"
            case cls.IMAGE_JPEG:
                return "jpeg"
            case cls.IMAGE_BMP:
                return "bmp"
            case cls.IMAGE_GIF:
                return "gif"
            case _:
                return ""
        
        
class DateUtil:
    
    YYYY = "%Y"

    YYYY_MM = "%Y-%m"

    YYYY_MM_DD = "%Y-%m-%d"

    YYYYMMDDHHMMSS = "%Y%m%d%H%M%S"

    YYYY_MM_DD_HH_MM_SS = "%Y-%m-%d %H:%M:%S"
    
    @classmethod
    def get_date_now(cls) -> str:
        """
        获取当前日期 %Y-%m-%d

        Returns:
            str: 当前日期
        """
        return datetime.now().strftime(cls.YYYY_MM_DD)
    
    @classmethod
    def get_datetime_now(cls,fmt=None) -> str:
        """
        获取当前日期 %Y%m%d%H%M%S

        Returns:
            str: 当前日期
        """
        fmt = fmt or cls.YYYYMMDDHHMMSS
        return datetime.now().strftime(fmt)
    
    @classmethod
    def get_time_now(cls) -> str:
        """
        获取当前日期 '%Y-%m-%d %H:%M:%S'

        Returns:
            str: 当前日期
        """
        return datetime.now().strftime(cls.YYYY_MM_DD_HH_MM_SS)
    
    @classmethod
    def get_date_path(cls) -> str:
        """
        获取当前日期 %Y/%m/%d

        Returns:
            str: 当前日期
        """
        return datetime.now().strftime("%Y/%m/%d")
    
    @classmethod
    def get_datepath(cls) -> str:
        """
        获取当前日期 %Y%m%d

        Returns:
            str: 当前日期
        """
        return datetime.now().strftime("%Y%m%d")


class FileUploadUtil:
    
    DEFAULT_MAX_SIZE = 50 * 1024 * 1024

    DEFAULT_FILE_NAME_LENGTH = 100
        
    @classmethod
    def upload(cls, file:FileStorage, base_path:str) -> str:
        '''
        上传文件
        
        Args:
            file(FileStorage): 文件对象
            base_path(str): 上传路径
        
        Returns:
            str: 资源路径
        '''
        fn_len = len(file.filename)
        if fn_len > cls.DEFAULT_FILE_NAME_LENGTH:
            raise Exception("文件名长度超过限制")
        cls.check_allowed(file, MimeTypeUtil.DEFAULT_ALLOWED_EXTENSION)
        filename = cls.extract_file_name(file)
        filepath = os.path.join(base_path, filename)
        file_parpath = os.path.dirname(filepath)
        if not os.path.exists(file_parpath):
            os.makedirs(file_parpath)
        file.save(filepath)
        resource_path = Constants.RESOURCE_PREFIX + "/" + filepath
        return resource_path
        
    @classmethod
    def check_allowed(cls, file:FileStorage, allowed_extensions:List[str]):
        '''
        文件大小校验
        
        Args:
            file(FileStorage): 文件对象
            allowed_extensions(List[str]): 允许的扩展名列表
        '''
        file_size = file.stream.tell()
        if file_size > cls.DEFAULT_MAX_SIZE:
            raise Exception("文件大小超过限制")
        extension = MimeTypeUtil.get_extension(file.content_type.lower())
        if extension not in allowed_extensions:
            if allowed_extensions == MimeTypeUtil.IMAGE_EXTENSION:
                raise Exception("图片格式不支持")
            elif allowed_extensions == MimeTypeUtil.VIDEO_EXTENSION:
                raise Exception("视频格式不支持")
            elif allowed_extensions == MimeTypeUtil.MEDIA_EXTENSION:
                raise Exception("媒体格式不支持")
            elif allowed_extensions == MimeTypeUtil.FLASH_EXTENSION:
                raise Exception("FLASH格式不支持")
            else:
                raise Exception("文件格式不支持")
            
    @classmethod
    def extract_file_name(cls, file:FileStorage) -> str:
        '''
        提取文件名
        
        Args:
            file(FileStorage): 文件对象
        
        Returns:
            str: 文件名
        '''
        "{}/{}_{}.{}".format(
            DateUtil.get_date_path(),
            os.path.basename(file.filename),
            Seq.get_seq_id(cls.upload_seq_type),
            cls.get_extension(file)
        )

    @classmethod
    def get_extension(cls, file:FileStorage) -> str:
        '''
        获取文件扩展名
        
        Args:
            file(FileStorage): 文件对象
        
        Returns:
            str: 文件扩展名
        '''
        extension = os.path.splitext(file.filename)[1]
        if extension is None or len(extension) == 0:
            extension = MimeTypeUtil.get_extension(file.content_type.lower())
        return extension

    @classmethod
    def get_filename(cls, filename:str) -> str:
        '''
        获取文件名
        
        Args:
            filename(str): 带路径和后缀的文件名
        
        Returns:
            str: 文件名
        '''
        return os.path.basename(filename)
    

class AtomicInteger:
    
    def __init__(self, initial=0):
        self._value = initial
        self._lock = threading.Lock()

    def get(self):
        '''
        获取当前值
        
        Returns:
            int: 当前值
        '''
        with self._lock:
            return self._value

    def set(self, value):
        '''
        设置新的值
        
        Args:
            value(int): 新的值
        '''
        with self._lock:
            self._value = value

    def increment(self):
        '''
        原子增加操作
        
        Returns:
            int: 新值
        '''
        with self._lock:
            self._value += 1
            return self._value

    def decrement(self):
        '''
        原子减少操作
        
        Returns:
            int: 新值
        '''
        with self._lock:
            self._value -= 1
            return self._value
        

class Seq:
    
    common_seq_type = "common"
    
    upload_seq_type = "upload"
    
    common_seq = AtomicInteger(1)
    
    upload_seq = AtomicInteger(1)
    
    matchine_code = "A"
    
    @classmethod
    def get_seq_id(cls, seq_name:str = "common") -> int:
        '''
        获取序列号
        
        Args:
            seq_name(str): 序列名称，common或upload
        
        Returns:
            int: 序列号
        '''
        ato = cls.upload_seq if seq_name == cls.upload_seq_type else cls.common_seq
        out = DateUtil.get_datetime_now() + cls.matchine_code + cls.get_seq(ato, 3)
        return out
    
    @classmethod
    def get_seq(cls, ato:AtomicInteger, length:int) -> str:
        '''
        获取指定长度的序列号
        
        Args:
            ato(AtomicInteger): 原子整数
            length(int): 序列号长度
        
        Returns:
            str: 序列号
        '''
        seq = str(ato.increment())
        if ato.get() > math.pow(10, length):
            ato.set(1)
        return StringUtil.left_pad(seq, length)
        

class MessageUtil:
    
    @staticmethod
    def message(code:str) -> str:
        """
        根据code获取消息

        Args:
            code (str): 消息代码

        Returns:
            str: 消息内容
        """
        # todo
        return code
        
    
class FileUtil:
    
    def delete_file(file_path:str) -> bool:
        '''
        删除文件
        
        Args:
            file_path(str): 文件路径
        
        Returns:
            bool: 是否成功
        '''
        flag = False
        if os.path.isfile(file_path) and os.path.exists(file_path):
            os.remove(file_path)
            flag = True
        return flag  


class DescriptUtil:
    
    @classmethod
    def get_raw(cls, func:Callable) -> Callable:
        """
        获取原始函数
        
        Args:
            func(Callable): 被装饰函数
        
        Returns:
            Callable: 原始函数
        """
        if hasattr(func, "__wrapped__"):
            func = func.__wrapped__
            return cls.get_raw(func)
        else:
            return func


class ExcelUtil:
    
    default_header_fill = {
        "start_color": "FFFFFFFF",
        "end_color": "FFFFFFFF",
        "fill_type": None, # "solid" or None
    }

    default_row_fill = {
        "start_color": "FFFFFFFF",
        "end_color": "FFFFFFFF",
        "fill_type": None, # "solid" or None
    }
    
    def __init__(self, model:Type[BaseModel]):
        self.model = model
    
    def write(self, data:List[BaseModel], sheetname:str) -> BytesIO:
        """
        写入Excel文件
        
        Args:
            data(List[BaseModel]): 数据
            sheetname(str): 工作表名
        
        Returns:
            BytesIO: 文件字节流
        """
        if len(data) == 0:
            raise NotFound(description="无法导出excel,数据为空")  
        workbook = Workbook()
        worksheet = workbook.create_sheet(title=sheetname)
        workbook.active = worksheet
        self.render_data(worksheet,data)
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def render_header(self, sheet:Worksheet,fill:PatternFill=None):
        """
        渲染Excel表头

        Args:
            sheet (Worksheet): 工作表
            fill(PatternFill): 表头填充
        """
        for col_index,access in enumerate(
            self.model.generate_excel_schema(),
            start=1
        ):
            _, access = access
            cell = sheet.cell(row=1,column=col_index,value=access.name)
            cell.fill = fill
            cell.font = access.header_font
    
    def render_row(self, sheet:Worksheet,row:BaseModel,row_index:int):
        """
        渲染Excel行数据
                
        Args:
            sheet (Worksheet): 工作表
            row (BaseModel): 行数据模型
            row_index(int): 行索引
        """
        default_row_fill = PatternFill(
            **self.default_row_fill
        )
        for col_index,access in enumerate(row.generate_excel_data(),start=1):
            _,access = access
            cell = sheet.cell(row=row_index,column=col_index,value=access.val)
            cell.alignment = access.alignment
            cell.fill = access.fill if access.fill else default_row_fill
            cell.font = access.row_font

    def render_footer(self, sheet:Worksheet):
        """
        渲染Excel表尾

        Args:
            sheet (Worksheet): 工作表
        """
        pass
    
    def render_data(self, sheet:Worksheet, data:List[BaseModel],header_fill:PatternFill=None):
        """
        渲染Excel数据
        
        Args:
            sheet (Worksheet): 工作表
            data(List[BaseModel]): 数据模型列表
        """
        if not header_fill:
            header_fill = PatternFill(
                **self.default_header_fill
            )
        
        self.render_header(sheet,header_fill)
        for row_index,row in enumerate(data,start=2):
            self.render_row(sheet,row,row_index)
        self.render_footer(sheet)
        
    def export_response(self, data:List[BaseModel], sheetname:str) -> Response:
        """
        响应Excel文件
        
        Args:
            data(List[BaseModel]): 数据
            sheetname(str): 工作表名
        
        Returns:
            Response: 文件流响应
        """
        output:BytesIO = self.write(data,sheetname)
        response = Response(
            response=output.getvalue(),
            status=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={time.time()}.xlsx"}
        )
        return response

    def import_template_response(self, sheetname:str) -> Response:
        """
        响应导入模板
        
        Args:
            sheetname(str): 工作表名
        
        Returns:
            Response: 文件流响应
        """
        output:BytesIO = self.write_template(sheetname)
        response = Response(
            response=output.getvalue(),
            status=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={self.model.__name__}_import_template.xlsx"}
        )
        return response
    
    def write_template(self,sheetname:str) -> BytesIO:
        """
        写入导入模板
        
        Args:
            sheetname(str): 工作表名
        
        Returns:
            BytesIO: 文件字节流
        """
        workbook = Workbook()
        worksheet = workbook.create_sheet(title=sheetname)
        workbook.active = worksheet
        
        self.render_template(worksheet)
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def render_template(self, sheet:Worksheet):
        """
        渲染导入模板
        
        Args:
            sheet (Worksheet): 工作表
        """
        header_fill = PatternFill(
            **self.default_header_fill
        )
        self.render_header(sheet,header_fill)
    
    def import_data(self, dto):
        """
        导入数据
        
        Args:
            dto(ImportDto): 导入数据模型
        
        Returns:
            List[BaseModel]: 导入数据模型列表
        """
        pass


def get_final_type(annotation) -> Type:
    args = get_args(annotation)
    if isinstance(args, tuple) and len(args) > 1:
        arg0 = args[0]
        if isinstance(arg0, Type):
            return arg0
        else:
            return get_final_type(arg0)


def get_final_model(annotation) -> Optional[type]:
    origin = get_origin(annotation)
    if origin is None:
        return annotation
    else:
        args = get_args(annotation)
        for arg in args:
            return get_final_model(arg)
        else:
            return None
